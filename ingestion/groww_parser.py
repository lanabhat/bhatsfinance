"""
Parser for Groww Excel portfolio exports.

Groww exports a single .xlsx file containing two sections:
  1. Stocks — header: "Stock Name, ISIN, Quantity, ..."
  2. Mutual Funds — header: "Scheme Name, AMC, Category, ..."

Each section is preceded by metadata rows (name, PAN, summary) that must be skipped.
The investor name and valuation date are extracted from those metadata rows.
"""
from datetime import date, datetime
from io import BytesIO


def _cell_str(cell) -> str:
    """Return cell value as a stripped string, or empty string."""
    if cell is None or cell.value is None:
        return ''
    return str(cell.value).strip()


def _row_values(row) -> list:
    return [_cell_str(c) for c in row]


def parse_groww_excel(file_bytes: bytes) -> dict:
    """
    Parse a Groww holdings Excel export.

    Returns:
        {
            "investor_name": str,
            "valuation_date": date | None,
            "stocks": [
                {
                    "name": str, "isin": str, "quantity": str,
                    "avg_buy_price": str, "buy_value": str,
                    "closing_price": str, "closing_value": str,
                }
            ],
            "mutual_funds": [
                {
                    "scheme_name": str, "amc": str, "category": str,
                    "sub_category": str, "folio_no": str, "units": str,
                    "invested_value": str, "current_value": str,
                }
            ],
        }
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("openpyxl is required for Groww Excel parsing")

    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    # Fallback: some xlsx files aren't readable in read_only mode
    _test_rows = list(wb.worksheets[0].rows)
    wb.close()
    if len(_test_rows) <= 1:
        wb = load_workbook(BytesIO(file_bytes), data_only=True)
    else:
        wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)

    investor_name = ''
    valuation_date = None
    stocks = []
    mutual_funds = []

    for sheet in wb.worksheets:
        all_rows = list(sheet.rows)
        i = 0
        n = len(all_rows)

        while i < n:
            vals = _row_values(all_rows[i])
            first = vals[0] if vals else ''

            # Extract investor name (appears in both sheets as "Name" row)
            if first == 'Name' and len(vals) > 1 and vals[1] and not investor_name:
                investor_name = vals[1]

            # Extract valuation date from "Holdings statement for stocks as on DD-MM-YYYY"
            # or "HOLDINGS AS ON YYYY-MM-DD" style rows
            if not valuation_date and first:
                low = first.lower()
                if 'holdings' in low and ('as on' in low or 'as of' in low):
                    # Try to find a date in the cell text itself
                    valuation_date = _extract_date_from_text(first)
                    # Also check next cell
                    if not valuation_date and len(vals) > 1 and vals[1]:
                        valuation_date = _extract_date_from_text(vals[1])

            # Detect stocks section header
            if first == 'Stock Name' and len(vals) > 1 and vals[1] == 'ISIN':
                headers = vals
                i += 1
                consecutive_blank = 0
                while i < n:
                    row_vals = _row_values(all_rows[i])
                    if not any(row_vals):
                        consecutive_blank += 1
                        # Two consecutive blank rows = end of section
                        if consecutive_blank >= 2:
                            break
                        i += 1
                        continue
                    consecutive_blank = 0
                    row_dict = dict(zip(headers, row_vals))
                    name = row_dict.get('Stock Name', '')
                    if name and name != 'Stock Name':
                        stocks.append({
                            'name': name,
                            'isin': row_dict.get('ISIN', ''),
                            'quantity': row_dict.get('Quantity', '0'),
                            'avg_buy_price': row_dict.get('Average buy price', '0'),
                            'buy_value': row_dict.get('Buy value', '0'),
                            'closing_price': row_dict.get('Closing price', '0'),
                            'closing_value': row_dict.get('Closing value', '0'),
                        })
                    i += 1
                continue

            # Detect MF section header
            if first == 'Scheme Name' and len(vals) > 1 and vals[1] == 'AMC':
                headers = vals
                i += 1
                consecutive_blank = 0
                while i < n:
                    row_vals = _row_values(all_rows[i])
                    if not any(row_vals):
                        consecutive_blank += 1
                        # Two consecutive blank rows = end of section
                        if consecutive_blank >= 2:
                            break
                        i += 1
                        continue
                    consecutive_blank = 0
                    row_dict = dict(zip(headers, row_vals))
                    name = row_dict.get('Scheme Name', '')
                    if name and name != 'Scheme Name':
                        mutual_funds.append({
                            'scheme_name': name,
                            'amc': row_dict.get('AMC', ''),
                            'category': row_dict.get('Category', ''),
                            'sub_category': row_dict.get('Sub-category', ''),
                            'folio_no': row_dict.get('Folio No.', ''),
                            'units': row_dict.get('Units', '0'),
                            'invested_value': row_dict.get('Invested Value', '0'),
                            'current_value': row_dict.get('Current Value', '0'),
                        })
                    i += 1
                continue

            i += 1

    wb.close()

    return {
        'investor_name': investor_name,
        'valuation_date': valuation_date or date.today(),
        'stocks': stocks,
        'mutual_funds': mutual_funds,
    }


def _extract_date_from_text(text: str):
    """Try to extract a date from a string like 'Holdings as on 17-06-2026' or 'HOLDINGS AS ON 2026-06-18'."""
    import re
    # Find date-like patterns in text
    patterns = [
        (r'(\d{4}-\d{2}-\d{2})', '%Y-%m-%d'),
        (r'(\d{2}-\d{2}-\d{4})', '%d-%m-%Y'),
        (r'(\d{2}/\d{2}/\d{4})', '%d/%m/%Y'),
    ]
    for pattern, fmt in patterns:
        m = re.search(pattern, text)
        if m:
            try:
                return datetime.strptime(m.group(1), fmt).date()
            except ValueError:
                continue
    return None
