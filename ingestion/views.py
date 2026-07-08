import difflib

from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView

from ingestion.serializers import CSVImportRequestSerializer
from ingestion.services import parse_and_process


class CSVImportView(APIView):
    def post(self, request):
        serializer = CSVImportRequestSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data
        result = parse_and_process(
            household_id=data['household_id'],
            filename=data.get('filename', ''),
            csv_content=data.get('csv_content'),
            rows=data.get('rows'),
        )
        return Response(result, status=status.HTTP_201_CREATED)


class ImportPreviewView(APIView):
    """Step 1: parse an uploaded file and return columns + first 5 preview rows."""

    def post(self, request):
        from ingestion.file_parser import parse_uploaded_file

        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'error': 'No file provided'}, status=400)

        file_type = request.data.get('file_type', '').lower()
        filename = file_obj.name or ''
        # auto-detect from extension if not supplied
        if not file_type and '.' in filename:
            ext = filename.rsplit('.', 1)[-1].lower()
            file_type = {'xlsx': 'excel', 'xls': 'excel', 'csv': 'csv',
                         'pdf': 'pdf', 'json': 'json'}.get(ext, ext)

        try:
            result = parse_uploaded_file(file_obj.read(), filename, file_type)
        except Exception as e:
            return Response({'error': str(e)}, status=400)

        return Response({
            'columns': result['columns'],
            'preview': result['rows'][:5],
            'total_rows': len(result['rows']),
            'all_rows': result['rows'],
        })


class ImportSchemasView(APIView):
    """Return field schemas for all import types (used by the frontend mapping UI)."""

    def get(self, request):
        from ingestion.import_schemas import IMPORT_SCHEMAS
        return Response(IMPORT_SCHEMAS)


class ImportApplyView(APIView):
    """Step 3: apply mapping + defaults and create model records."""

    def post(self, request):
        from ingestion.universal_importer import apply_import

        data = request.data
        household_id = data.get('household_id')
        import_type = data.get('import_type')
        rows = data.get('rows', [])
        mapping = data.get('mapping', {})
        defaults = data.get('defaults', {})

        if not household_id:
            return Response({'error': 'household_id is required'}, status=400)
        if not import_type:
            return Response({'error': 'import_type is required'}, status=400)
        if not rows:
            return Response({'error': 'rows is required'}, status=400)

        try:
            result = apply_import(
                household_id=int(household_id),
                import_type=import_type,
                rows=rows,
                mapping=mapping,
                defaults=defaults,
            )
        except Exception as e:
            return Response({'error': str(e)}, status=400)

        return Response(result, status=201)


def _detect_and_parse(file_bytes: bytes, filename: str) -> tuple[dict, str]:
    """
    Auto-detect whether a file is a Groww or Upstox export and parse it.
    Returns (parsed_dict, source) where source is 'groww' or 'upstox'.
    Detection order:
      1. Filename contains 'upstox' (case-insensitive)
      2. First 10 rows contain 'UPSTOX' or 'UCC' anywhere in any cell value
      3. Fall back to Groww parser
    """
    from ingestion.groww_parser import parse_groww_excel
    from ingestion.upstox_parser import parse_upstox_excel
    from openpyxl import load_workbook
    from io import BytesIO

    # 1. Filename-based detection
    import re as _re
    fname_lower = filename.lower()
    # Upstox filenames: 'holdings_18-06-2026_HT9515.xlsx' or contain 'upstox'/'rksv'
    _upstox_fname = (
        'upstox' in fname_lower
        or 'rksv' in fname_lower
        or bool(_re.match(r'holdings_\d{2}-\d{2}-\d{4}', fname_lower))
    )
    if _upstox_fname:
        return parse_upstox_excel(file_bytes), 'upstox'

    # 2. Content-based detection — scan all cell values in first 10 rows
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    sheet = wb.worksheets[0]
    rows_preview = list(sheet.rows)[:10]
    wb.close()
    # Fallback for files unreadable in read_only mode
    if len(rows_preview) <= 1:
        wb2 = load_workbook(BytesIO(file_bytes), data_only=True)
        sheet = wb2.worksheets[0]
        rows_preview = list(sheet.rows)[:10]
        wb2.close()
    is_upstox = False
    for row in rows_preview:
        for cell in row:
            val = str(cell.value or '').strip().upper()
            if 'UPSTOX' in val or 'RKSV' in val or val == 'UCC':
                is_upstox = True
                break
        if is_upstox:
            break

    if is_upstox:
        return parse_upstox_excel(file_bytes), 'upstox'
    return parse_groww_excel(file_bytes), 'groww'


def _fuzzy_match_member(investor_name: str, member_names: dict):
    """Returns (member, confidence) or (None, 0)."""
    if not investor_name or not member_names:
        return None, 0.0
    name_lower = investor_name.lower().strip()
    close = difflib.get_close_matches(name_lower, member_names.keys(), n=1, cutoff=0.4)
    if close:
        seq = difflib.SequenceMatcher(None, name_lower, close[0])
        return member_names[close[0]], round(seq.ratio(), 2)
    for key, m in member_names.items():
        if name_lower in key or key in name_lower:
            return m, 0.5
    return None, 0.0


class GrowwPreviewView(APIView):
    """
    Step 1: parse one or more uploaded Groww/Upstox Excel files and return
    metadata (investor name, holding counts) with fuzzy-matched member.
    Accepts both Groww and Upstox formats — auto-detected per file.
    """

    def post(self, request):
        from core.models import Household

        household_id = request.data.get('household_id')
        if not household_id:
            return Response({'error': 'household_id is required'}, status=400)

        try:
            household = Household.objects.get(pk=int(household_id))
        except Household.DoesNotExist:
            return Response({'error': 'Household not found'}, status=404)

        members = list(household.members.filter(is_active=True))
        member_names = {m.full_name.lower(): m for m in members}

        files = request.FILES.getlist('files')
        if not files:
            single = request.FILES.get('file')
            if single:
                files = [single]
        if not files:
            return Response({'error': 'No files provided'}, status=400)

        member_list = [{'id': m.id, 'name': m.full_name, 'relation': m.relation_type} for m in members]

        results = []
        for f in files:
            try:
                parsed, source = _detect_and_parse(f.read(), f.name)
            except Exception as e:
                results.append({'filename': f.name, 'error': str(e), 'members': member_list})
                continue

            investor_name = parsed.get('investor_name', '')
            matched_member, confidence = _fuzzy_match_member(investor_name, member_names)

            # Build summary counts based on source
            if source == 'groww':
                summary = {
                    'stocks_count': len(parsed.get('stocks', [])),
                    'mf_count': len(parsed.get('mutual_funds', [])),
                    'holdings_count': 0,
                }
            else:
                summary = {
                    'stocks_count': 0,
                    'mf_count': 0,
                    'holdings_count': len(parsed.get('holdings', [])),
                }

            results.append({
                'filename': f.name,
                'source': source,
                'investor_name': investor_name,
                'valuation_date': str(parsed.get('valuation_date', '')),
                **summary,
                'matched_member': {
                    'id': matched_member.id,
                    'name': matched_member.full_name,
                    'relation': matched_member.relation_type,
                    'confidence': confidence,
                } if matched_member else None,
                'members': [{'id': m.id, 'name': m.full_name} for m in members],
            })

        return Response(results)


class GrowwApplyView(APIView):
    """
    Step 2: apply confirmed member assignments and import holdings.
    Handles both Groww and Upstox formats — auto-detected per file.
    """

    def post(self, request):
        from ingestion.universal_importer import apply_groww_import, apply_upstox_import
        from core.models import Household, Member
        import json

        household_id = request.data.get('household_id')
        if not household_id:
            return Response({'error': 'household_id is required'}, status=400)

        try:
            household = Household.objects.get(pk=int(household_id))
        except Household.DoesNotExist:
            return Response({'error': 'Household not found'}, status=404)

        raw_assignments = request.data.get('assignments', '[]')
        try:
            assignments = json.loads(raw_assignments) if isinstance(raw_assignments, str) else raw_assignments
        except Exception:
            return Response({'error': 'assignments must be a JSON array'}, status=400)

        assignment_map = {a['filename']: int(a['member_id']) for a in assignments if a.get('filename') and a.get('member_id')}

        files = request.FILES.getlist('files')
        if not files:
            single = request.FILES.get('file')
            if single:
                files = [single]
        if not files:
            return Response({'error': 'No files provided'}, status=400)

        all_results = []
        for f in files:
            member_id = assignment_map.get(f.name)
            if not member_id:
                all_results.append({'filename': f.name, 'error': 'No member assigned'})
                continue

            try:
                member = Member.objects.get(pk=member_id, household=household)
            except Member.DoesNotExist:
                all_results.append({'filename': f.name, 'error': f'Member {member_id} not found'})
                continue

            try:
                parsed, source = _detect_and_parse(f.read(), f.name)
                if source == 'upstox':
                    result = apply_upstox_import(household, member, parsed)
                    result['holdings_parsed'] = len(parsed.get('holdings', []))
                else:
                    result = apply_groww_import(household, member, parsed)
                    result['stocks_parsed'] = len(parsed.get('stocks', []))
                    result['mf_parsed'] = len(parsed.get('mutual_funds', []))
                result['filename'] = f.name
                result['source'] = source
                result['member_name'] = member.full_name
                result['investor_name'] = parsed.get('investor_name', '')
                all_results.append(result)
            except Exception as e:
                all_results.append({'filename': f.name, 'member_name': member.full_name, 'error': str(e)})


class FDAdvicePreviewView(APIView):
    """
    Step 1: decrypt + parse one or more FD advice / RD statement PDFs and
    return extracted, user-editable fields per file, plus a fuzzy-matched
    member suggestion. Saved household passwords are tried automatically
    before falling back to any explicitly supplied password.
    """

    def post(self, request):
        import json

        from core.models import Household
        from ingestion.fd_parser import FDParseError, parse_deposit_text
        from ingestion.models import SavedPdfPassword
        from ingestion.pdf_decrypt import IncorrectPasswordError, decrypt_and_extract_text

        household_id = request.data.get('household_id')
        if not household_id:
            return Response({'error': 'household_id is required'}, status=400)

        try:
            household = Household.objects.get(pk=int(household_id))
        except Household.DoesNotExist:
            return Response({'error': 'Household not found'}, status=404)

        try:
            passwords = json.loads(request.data.get('passwords', '{}'))
        except Exception:
            passwords = {}
        try:
            save_passwords = json.loads(request.data.get('save_passwords', '{}'))
        except Exception:
            save_passwords = {}

        members = household.members.filter(is_active=True)
        member_names = {m.full_name.lower(): m for m in members}
        member_list = [{'id': m.id, 'name': m.full_name, 'relation': m.relation_type} for m in members]

        saved_passwords = list(
            SavedPdfPassword.objects.filter(household=household).order_by('-last_used_at', '-created_at')
        )

        files = request.FILES.getlist('files')
        if not files:
            single = request.FILES.get('file')
            if single:
                files = [single]
        if not files:
            return Response({'error': 'No files provided'}, status=400)

        results = []
        for f in files:
            file_bytes = f.read()

            candidates: list[tuple[str, object]] = []
            supplied = passwords.get(f.name)
            if supplied is not None:
                candidates.append((supplied, None))
            for saved in saved_passwords:
                candidates.append((saved.password, saved))
            if not candidates:
                candidates.append(('', None))

            text = None
            matched_saved = None
            winning_password = None
            password_error = None
            corrupt_error = None
            for pw, saved_obj in candidates:
                try:
                    text = decrypt_and_extract_text(file_bytes, pw)
                    matched_saved = saved_obj
                    winning_password = pw
                    break
                except IncorrectPasswordError as e:
                    password_error = e
                    continue
                except ValueError as e:
                    corrupt_error = e
                    break

            if text is None:
                if corrupt_error is not None:
                    results.append({'filename': f.name, 'error': str(corrupt_error)})
                else:
                    results.append({'filename': f.name, 'error': str(password_error) if password_error else 'Could not read PDF.', 'error_code': 'bad_password'})
                continue

            if matched_saved is not None:
                from django.utils import timezone
                matched_saved.last_used_at = timezone.now()
                matched_saved.save(update_fields=['last_used_at'])
            elif supplied is not None and save_passwords.get(f.name, True):
                from django.utils import timezone
                already_saved = any(s.password == winning_password for s in saved_passwords)
                if not already_saved:
                    new_saved = SavedPdfPassword.objects.create(
                        household=household,
                        password=winning_password,
                        last_used_at=timezone.now(),
                    )
                    saved_passwords.insert(0, new_saved)

            try:
                parsed = parse_deposit_text(text)
            except FDParseError as e:
                results.append({'filename': f.name, 'error': str(e)})
                continue

            matched_member, confidence = _fuzzy_match_member(parsed.get('member_name_raw', ''), member_names)

            results.append({
                'filename': f.name,
                **parsed,
                'matched_member': (
                    {'id': matched_member.id, 'name': matched_member.full_name,
                     'relation': matched_member.relation_type, 'confidence': confidence}
                    if matched_member else None
                ),
                'members': member_list,
            })

        return Response(results)


class FDAdviceApplyView(APIView):
    """Step 2: commit records from user-confirmed/edited FD/RD field values."""

    def post(self, request):
        from ingestion.universal_importer import apply_fd_advice_import, apply_rd_statement_import
        from core.models import Household, Member
        from instruments.models import Account

        household_id = request.data.get('household_id')
        if not household_id:
            return Response({'error': 'household_id is required'}, status=400)

        try:
            household = Household.objects.get(pk=int(household_id))
        except Household.DoesNotExist:
            return Response({'error': 'Household not found'}, status=404)

        items = request.data.get('items', [])
        if not items:
            return Response({'error': 'items is required'}, status=400)

        all_results = []
        for item in items:
            member = None
            member_id = item.get('member_id')
            if member_id:
                try:
                    member = Member.objects.get(pk=member_id, household=household)
                except Member.DoesNotExist:
                    pass

            try:
                if item.get('doc_type') == 'rd_statement':
                    account_id = item.get('account_id')
                    if not account_id:
                        raise ValueError('account_id is required for RD statement imports (select the account future installments will be paid from).')
                    try:
                        account = Account.objects.get(pk=account_id, household=household)
                    except Account.DoesNotExist:
                        raise ValueError(f'Account {account_id} not found')
                    result = apply_rd_statement_import(household, member, item, account)
                else:
                    result = apply_fd_advice_import(household, member, item)
                result['filename'] = item.get('filename', '')
                all_results.append(result)
            except Exception as e:
                all_results.append({'filename': item.get('filename', ''), 'error': str(e)})

        return Response(all_results, status=201)


class NpsPreviewView(APIView):
    """Step 1: parse one or more uploaded NPS transaction statement CSVs (Tier I/II)."""

    def post(self, request):
        from core.models import Household
        from ingestion.nps_parser import NpsParseError, parse_nps_statement

        household_id = request.data.get('household_id')
        if not household_id:
            return Response({'error': 'household_id is required'}, status=400)

        try:
            household = Household.objects.get(pk=int(household_id))
        except Household.DoesNotExist:
            return Response({'error': 'Household not found'}, status=404)

        members = household.members.filter(is_active=True)
        member_names = {m.full_name.lower(): m for m in members}
        member_list = [{'id': m.id, 'name': m.full_name, 'relation': m.relation_type} for m in members]

        files = request.FILES.getlist('files')
        if not files:
            single = request.FILES.get('file')
            if single:
                files = [single]
        if not files:
            return Response({'error': 'No files provided'}, status=400)

        results = []
        for f in files:
            try:
                text = f.read().decode('utf-8-sig')
            except UnicodeDecodeError:
                results.append({'filename': f.name, 'error': 'Could not read file as UTF-8 text.'})
                continue

            try:
                parsed = parse_nps_statement(text)
            except NpsParseError as e:
                results.append({'filename': f.name, 'error': str(e)})
                continue

            matched_member, confidence = _fuzzy_match_member(parsed.get('subscriber_name', ''), member_names)

            results.append({
                'filename': f.name,
                **parsed,
                'matched_member': (
                    {'id': matched_member.id, 'name': matched_member.full_name,
                     'relation': matched_member.relation_type, 'confidence': confidence}
                    if matched_member else None
                ),
                'members': member_list,
            })

        return Response(results)


class NpsApplyView(APIView):
    """Step 2: commit records from user-confirmed NPS statement parse results."""

    def post(self, request):
        from core.models import Household, Member
        from ingestion.universal_importer import apply_nps_statement_import
        from instruments.models import Account

        household_id = request.data.get('household_id')
        if not household_id:
            return Response({'error': 'household_id is required'}, status=400)

        try:
            household = Household.objects.get(pk=int(household_id))
        except Household.DoesNotExist:
            return Response({'error': 'Household not found'}, status=404)

        items = request.data.get('items', [])
        if not items:
            return Response({'error': 'items is required'}, status=400)

        all_results = []
        for item in items:
            member = None
            member_id = item.get('member_id')
            if member_id:
                try:
                    member = Member.objects.get(pk=member_id, household=household)
                except Member.DoesNotExist:
                    pass

            account = None
            account_id = item.get('account_id')
            if account_id:
                try:
                    account = Account.objects.get(pk=account_id, household=household)
                except Account.DoesNotExist:
                    pass

            try:
                result = apply_nps_statement_import(
                    household, member, item,
                    account=account,
                    affects_balance=bool(item.get('affects_balance', False)),
                )
                result['filename'] = item.get('filename', '')
                all_results.append(result)
            except Exception as e:
                all_results.append({'filename': item.get('filename', ''), 'error': str(e)})

        return Response(all_results, status=201)
