"""
Parser for Upstox holdings Excel exports.

Format:
  - Metadata rows at top (company name, UCC, Name, PAN, Report Date, Generated On)
  - Blank row
  - Header row: ISIN, Scrip Name, Free Qty, Current Qty, Freeze Qty, Locked Qty,
                Pledge Qty, Value Date, Rate, Valuation
  - Data rows until blank row
  - Footer note at bottom (ignored)

Investor name extracted from "Name" metadata row.
Valuation date from "Report Date" or "Value Date" column in data rows.
"""
import re
from datetime import date, datetime
from io import BytesIO


def _cell_str(cell) -> str:
    if cell is None or cell.value is None:
        return ''
    return str(cell.value).strip()


def _row_values(row) -> list:
    return [_cell_str(c) for c in row]


def _parse_date(val: str) -> date | None:
    for fmt in ('%d-%m-%Y', '%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y'):
        try:
            return datetime.strptime(val.strip(), fmt).date()
        except ValueError:
            continue
    return None


def parse_upstox_excel(file_bytes: bytes) -> dict:
    """
    Parse an Upstox holdings Excel export.

    Returns:
        {
            "investor_name": str,
            "valuation_date": date,
            "holdings": [
                {
                    "isin": str,
                    "name": str,
                    "quantity": str,
                    "rate": str,        # closing price per unit
                    "valuation": str,   # total market value
                    "value_date": str,
                }
            ]
        }
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("openpyxl is required for Upstox Excel parsing")

    # Try read_only first; fall back to full load if it yields ≤1 row
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    sheet = wb.worksheets[0]
    all_rows = list(sheet.rows)
    wb.close()
    if len(all_rows) <= 1:
        wb2 = load_workbook(BytesIO(file_bytes), data_only=True)
        sheet = wb2.worksheets[0]
        all_rows = list(sheet.rows)
        wb2.close()

    investor_name = ''
    report_date = None
    holdings = []

    n = len(all_rows)
    i = 0

    while i < n:
        vals = _row_values(all_rows[i])
        first = vals[0] if vals else ''

        # Extract investor name — check first non-empty cell after 'Name'
        if first == 'Name':
            for v in vals[1:]:
                if v:
                    investor_name = v
                    break

        # Extract report date — check first non-empty cell after 'Report Date'
        if first == 'Report Date':
            for v in vals[1:]:
                if v:
                    report_date = _parse_date(v)
                    break

        # Detect data header row by ISIN + Scrip Name
        if first == 'ISIN' and len(vals) > 1 and 'Scrip' in vals[1]:
            # Build column index map from headers
            headers = vals
            col = {name: idx for idx, name in enumerate(headers)}
            i += 1
            consecutive_blank = 0
            while i < n:
                row_vals = _row_values(all_rows[i])
                # Pad row to header length so index access is safe
                while len(row_vals) < len(headers):
                    row_vals.append('')
                if not any(row_vals):
                    consecutive_blank += 1
                    if consecutive_blank >= 2:
                        break
                    i += 1
                    continue
                consecutive_blank = 0
                first_cell = row_vals[0]
                # ISINs are alphanumeric, 12 chars — stop at footer text rows
                if first_cell and not re.match(r'^[A-Z0-9]{12}$', first_cell):
                    break

                def _get(col_name, default='0'):
                    idx = col.get(col_name)
                    if idx is None:
                        return default
                    return row_vals[idx] if idx < len(row_vals) else default

                isin = _get('ISIN', '').strip()
                name = _get('Scrip Name', '').strip()
                if isin and name:
                    vd_str = _get('Value Date', '')
                    vd = _parse_date(vd_str) if vd_str else None
                    rate = _get('Rate', '0')
                    valuation = _get('Valuation', '0')
                    # Fallback: if Rate/Valuation missing, use last two numeric columns
                    if (not rate or rate == '0') and (not valuation or valuation == '0'):
                        numeric_vals = [v for v in row_vals if v and v != '0.0000']
                        try:
                            valuation = numeric_vals[-1] if numeric_vals else '0'
                            rate = numeric_vals[-2] if len(numeric_vals) >= 2 else '0'
                        except IndexError:
                            pass
                    holdings.append({
                        'isin': isin,
                        'name': name,
                        'quantity': _get('Current Qty', '0'),
                        'rate': rate,
                        'valuation': valuation,
                        'value_date': str(vd or report_date or date.today()),
                    })
                i += 1
            continue

        i += 1

    return {
        'investor_name': investor_name,
        'valuation_date': report_date or date.today(),
        'holdings': holdings,
    }
