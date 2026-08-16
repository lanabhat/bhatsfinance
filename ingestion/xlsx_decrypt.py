"""
Decrypt password-protected .xlsx files (OLE2/CFB container) into a plain
openpyxl Workbook.

msoffcrypto-tool decrypts the OLE2 container into an in-memory zip stream
first, then openpyxl reads that — same two-step shape as pdf_decrypt.py's
pikepdf -> pdfplumber handoff for encrypted PDFs.
"""
import io


class IncorrectPasswordError(Exception):
    pass


def decrypt_and_load_workbook(file_bytes: bytes, password: str | None, data_only: bool = True):
    import msoffcrypto
    import openpyxl
    from msoffcrypto.exceptions import DecryptionError, InvalidKeyError

    stream = io.BytesIO(file_bytes)

    try:
        office_file = msoffcrypto.OfficeFile(stream)
        is_encrypted = office_file.is_encrypted()
    except Exception:
        is_encrypted = False

    if not is_encrypted:
        try:
            return openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=data_only)
        except Exception as e:
            raise ValueError(f"Could not read Excel file: {e}")

    try:
        office_file.load_key(password=password or "")
        decrypted = io.BytesIO()
        office_file.decrypt(decrypted)
        decrypted.seek(0)
    except (InvalidKeyError, DecryptionError):
        raise IncorrectPasswordError("Incorrect password for this Excel file.")
    except Exception as e:
        raise ValueError(f"Could not read Excel file: {e}")

    try:
        return openpyxl.load_workbook(decrypted, data_only=data_only)
    except Exception as e:
        raise ValueError(f"Could not read Excel file: {e}")
