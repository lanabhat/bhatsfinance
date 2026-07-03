"""
Account statement report: reverses the current-balance formula to compute
an opening balance for a chosen date range and a user-editable transaction
selection, then renders the result as PDF or Excel for printing/sharing.

Known simplification: the opening balance is derived by walking the
account's *current* balance backward through only the selected
transactions (current_balance - selected_inflows + selected_outflows).
This is exactly correct when the selected transactions are contiguous with
"today" — i.e. there are no unselected or out-of-range transactions between
the report's end_date and today. If the end_date is in the past, or a
transaction dated after other selected ones is excluded, the computed
opening balance will be off by the effect of those uncounted transactions.
This mirrors the simple mental model the report is meant to support
(a account statement working back from the balance you see right now),
not a full historical ledger reconstruction.
"""
import io
from decimal import Decimal


def get_statement_data(household_id: int, account_ids: list[int], start_date, end_date) -> list[dict]:
    from instruments.models import Account
    from instruments.services import compute_account_balance
    from ledger.models import Transaction

    accounts = Account.objects.filter(household_id=household_id, id__in=account_ids)

    results = []
    for account in accounts:
        balance = compute_account_balance(account)
        txs = (
            Transaction.objects
            .filter(account=account, tx_date__gte=start_date, tx_date__lte=end_date)
            .order_by('tx_date', 'id')
        )
        transactions = [
            {
                'id': tx.id,
                'tx_date': tx.tx_date.isoformat(),
                'description': tx.description or tx.external_reference or tx.get_transaction_type_display(),
                'direction': tx.direction,
                'transaction_type': tx.transaction_type,
                'amount': str(tx.amount),
            }
            for tx in txs
        ]
        results.append({
            'account_id': account.id,
            'account_name': account.name,
            'institution_name': account.institution_name,
            'current_balance': str(balance['current_balance']),
            'transactions': transactions,
        })
    return results


def compute_opening_balance(current_balance: Decimal, transactions: list[dict], excluded_ids: set[int]) -> Decimal:
    """
    transactions: list of dicts with 'id', 'direction' ('inflow'/'outflow'), 'amount' (str/Decimal).
    excluded_ids: transaction ids to exclude from the selection.
    """
    selected = [t for t in transactions if t['id'] not in excluded_ids]
    inflow_sum = sum((Decimal(str(t['amount'])) for t in selected if t['direction'] == 'inflow'), Decimal('0'))
    outflow_sum = sum((Decimal(str(t['amount'])) for t in selected if t['direction'] == 'outflow'), Decimal('0'))
    return current_balance - inflow_sum + outflow_sum


def build_statement_report(
    household_id: int,
    account_ids: list[int],
    start_date,
    end_date,
    excluded_transaction_ids: dict[int, list[int]] | None = None,
    opening_balance_overrides: dict[int, str] | None = None,
) -> list[dict]:
    """
    Returns one dict per account with:
    {
        account_id, account_name, institution_name,
        opening_balance: Decimal, closing_balance: Decimal,
        rows: [{tx_date, description, direction, amount, running_balance}, ...]
    }
    """
    excluded_transaction_ids = excluded_transaction_ids or {}
    opening_balance_overrides = opening_balance_overrides or {}

    raw_accounts = get_statement_data(household_id, account_ids, start_date, end_date)

    report = []
    for acc in raw_accounts:
        account_id = acc['account_id']
        current_balance = Decimal(acc['current_balance'])
        excluded_ids = set(excluded_transaction_ids.get(account_id, []) or excluded_transaction_ids.get(str(account_id), []))
        selected = [t for t in acc['transactions'] if t['id'] not in excluded_ids]

        override_raw = opening_balance_overrides.get(account_id) or opening_balance_overrides.get(str(account_id))
        if override_raw not in (None, ''):
            opening_balance = Decimal(str(override_raw))
        else:
            opening_balance = compute_opening_balance(current_balance, acc['transactions'], excluded_ids)

        rows = []
        running = opening_balance
        for t in selected:
            amount = Decimal(str(t['amount']))
            if t['direction'] == 'inflow':
                running += amount
            else:
                running -= amount
            rows.append({
                'tx_date': t['tx_date'],
                'description': t['description'],
                'direction': t['direction'],
                'amount': amount,
                'running_balance': running,
            })

        closing_balance = running

        report.append({
            'account_id': account_id,
            'account_name': acc['account_name'],
            'institution_name': acc['institution_name'],
            'opening_balance': opening_balance,
            'closing_balance': closing_balance,
            'rows': rows,
        })

    return report


# ---------------------------------------------------------------------------
# PDF rendering
# ---------------------------------------------------------------------------

def render_statement_pdf(report_rows: list[dict], meta: dict) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        topMargin=20 * mm, bottomMargin=20 * mm, leftMargin=18 * mm, rightMargin=18 * mm,
    )
    styles = getSampleStyleSheet()
    cell_style = ParagraphStyle('TableCell', parent=styles['Normal'], fontSize=8, leading=10)

    story = []
    story.append(Paragraph(
        f"Generated on {meta.get('generated_on', '')} for {meta.get('household_name', '')}",
        ParagraphStyle('Meta', parent=styles['Normal'], fontSize=8, textColor=colors.grey),
    ))
    story.append(Spacer(1, 8))

    for i, acc in enumerate(report_rows):
        title = acc['account_name']
        if acc.get('institution_name'):
            title += f" ({acc['institution_name']})"
        story.append(Paragraph(title, styles['Heading2']))
        story.append(Paragraph(
            f"Statement period: {meta['start_date']} to {meta['end_date']}", styles['Normal'],
        ))
        story.append(Spacer(1, 4))
        story.append(Paragraph(f"<b>Opening Balance: Rs. {acc['opening_balance']:,.2f}</b>", styles['Normal']))
        story.append(Spacer(1, 6))

        header = ['Date', 'Description', 'Direction', 'Amount', 'Running Balance']
        data = [header]
        for row in acc['rows']:
            signed = f"{'+' if row['direction'] == 'inflow' else '-'}{row['amount']:,.2f}"
            data.append([
                row['tx_date'],
                Paragraph(row['description'], cell_style),
                row['direction'].capitalize(),
                signed,
                f"{row['running_balance']:,.2f}",
            ])

        col_widths = [55, 190, 55, 75, 78]
        table = Table(data, colWidths=[w for w in col_widths], repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f1f5f9')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
            ('ALIGN', (3, 0), (4, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        story.append(table)
        story.append(Spacer(1, 6))
        story.append(Paragraph(f"<b>Closing Balance: Rs. {acc['closing_balance']:,.2f}</b>", styles['Normal']))

        if i < len(report_rows) - 1:
            story.append(PageBreak())

    doc.build(story)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# Excel rendering
# ---------------------------------------------------------------------------

def render_statement_xlsx(report_rows: list[dict], meta: dict) -> bytes:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    bold = Font(bold=True)
    header_fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')

    for acc in report_rows:
        title = acc['account_name'][:31] or f"Account {acc['account_id']}"
        # Excel sheet titles must be unique and <=31 chars; disambiguate on collision.
        base_title, suffix = title, 1
        existing = set(wb.sheetnames)
        while title in existing:
            suffix += 1
            title = f"{base_title[:28]} ({suffix})"
        ws = wb.create_sheet(title=title)

        header_text = acc['account_name']
        if acc.get('institution_name'):
            header_text += f" ({acc['institution_name']})"
        ws.merge_cells('A1:E1')
        ws['A1'] = header_text
        ws['A1'].font = Font(bold=True, size=14)

        ws['A2'] = f"Statement period: {meta['start_date']} to {meta['end_date']}"

        ws['A4'] = 'Opening Balance'
        ws['A4'].font = bold
        ws['B4'] = float(acc['opening_balance'])
        ws['B4'].number_format = '#,##0.00'
        ws['B4'].font = bold

        header_row = 6
        headers = ['Date', 'Description', 'Direction', 'Amount', 'Running Balance']
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col, value=h)
            cell.font = bold
            cell.fill = header_fill

        row_idx = header_row + 1
        for row in acc['rows']:
            signed_amount = row['amount'] if row['direction'] == 'inflow' else -row['amount']
            ws.cell(row=row_idx, column=1, value=row['tx_date'])
            ws.cell(row=row_idx, column=2, value=row['description'])
            ws.cell(row=row_idx, column=3, value=row['direction'].capitalize())
            amt_cell = ws.cell(row=row_idx, column=4, value=float(signed_amount))
            amt_cell.number_format = '#,##0.00'
            bal_cell = ws.cell(row=row_idx, column=5, value=float(row['running_balance']))
            bal_cell.number_format = '#,##0.00'
            row_idx += 1

        ws.cell(row=row_idx + 1, column=1, value='Closing Balance').font = bold
        closing_cell = ws.cell(row=row_idx + 1, column=2, value=float(acc['closing_balance']))
        closing_cell.font = bold
        closing_cell.number_format = '#,##0.00'

        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 16
        ws.column_dimensions['E'].width = 18
        ws.print_area = f'A1:E{row_idx + 1}'
        ws.page_setup.orientation = 'portrait'
        ws.print_options.horizontalCentered = True

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
