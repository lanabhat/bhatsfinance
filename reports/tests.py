import io
from datetime import date
from decimal import Decimal

import openpyxl
from django.test import TestCase

from core.models import Household, Member
from instruments.models import Account
from ledger.models import Transaction
from reports.services import (
    build_statement_report,
    compute_opening_balance,
    get_statement_data,
    render_statement_pdf,
    render_statement_xlsx,
)


class OpeningBalanceReversalTests(TestCase):
    def setUp(self):
        self.household = Household.objects.create(name='Sharma Family')
        self.member = Member.objects.create(household=self.household, full_name='Amit Sharma')
        self.account = Account.objects.create(
            household=self.household,
            name='Shared Account',
            account_type=Account.AccountType.BANK,
            opening_balance=Decimal('0.00'),
        )

    def test_single_outflow_reverses_to_higher_opening_balance(self):
        # Confirmed example: current balance 100,000 with one selected 20,000
        # outflow in range -> opening balance should be 120,000.
        transactions = [{'id': 1, 'direction': 'outflow', 'amount': '20000.00'}]
        opening = compute_opening_balance(Decimal('100000'), transactions, excluded_ids=set())
        self.assertEqual(opening, Decimal('120000'))

    def test_no_selected_transactions_opening_equals_current(self):
        transactions = [{'id': 1, 'direction': 'outflow', 'amount': '20000.00'}]
        opening = compute_opening_balance(Decimal('100000'), transactions, excluded_ids={1})
        self.assertEqual(opening, Decimal('100000'))

    def test_mixed_inflow_and_outflow_selection(self):
        transactions = [
            {'id': 1, 'direction': 'outflow', 'amount': '20000.00'},
            {'id': 2, 'direction': 'inflow', 'amount': '5000.00'},
        ]
        # opening = current - inflow + outflow = 100000 - 5000 + 20000 = 115000
        opening = compute_opening_balance(Decimal('100000'), transactions, excluded_ids=set())
        self.assertEqual(opening, Decimal('115000'))


class StatementReportTests(TestCase):
    def setUp(self):
        self.household = Household.objects.create(name='Sharma Family')
        self.member = Member.objects.create(household=self.household, full_name='Amit Sharma')
        self.account_a = Account.objects.create(
            household=self.household, name='Uncle Fund', account_type=Account.AccountType.BANK,
            opening_balance=Decimal('0.00'),
        )
        self.account_b = Account.objects.create(
            household=self.household, name='Personal Savings', account_type=Account.AccountType.BANK,
            opening_balance=Decimal('0.00'),
        )
        # Account A: one inflow of 100,000 (uncle's deposit) inside the range.
        self.tx_a = Transaction.objects.create(
            household=self.household, account=self.account_a, member=self.member,
            tx_date=date(2026, 1, 10), amount=Decimal('100000.00'),
            direction=Transaction.Direction.INFLOW, transaction_type=Transaction.TransactionType.DEPOSIT,
            description="Uncle's deposit",
        )
        # Account B: unrelated transaction, must not leak into account A's report.
        self.tx_b = Transaction.objects.create(
            household=self.household, account=self.account_b, member=self.member,
            tx_date=date(2026, 1, 15), amount=Decimal('5000.00'),
            direction=Transaction.Direction.OUTFLOW, transaction_type=Transaction.TransactionType.WITHDRAWAL,
            description='Groceries',
        )

    def test_get_statement_data_scopes_transactions_per_account(self):
        data = get_statement_data(
            self.household.id, [self.account_a.id, self.account_b.id],
            date(2026, 1, 1), date(2026, 1, 31),
        )
        self.assertEqual(len(data), 2)
        by_id = {d['account_id']: d for d in data}
        self.assertEqual(len(by_id[self.account_a.id]['transactions']), 1)
        self.assertEqual(by_id[self.account_a.id]['transactions'][0]['description'], "Uncle's deposit")
        self.assertEqual(len(by_id[self.account_b.id]['transactions']), 1)
        self.assertEqual(by_id[self.account_b.id]['transactions'][0]['description'], 'Groceries')

    def test_build_statement_report_multi_account_independent_balances(self):
        report = build_statement_report(
            self.household.id, [self.account_a.id, self.account_b.id],
            date(2026, 1, 1), date(2026, 1, 31),
        )
        self.assertEqual(len(report), 2)
        by_id = {r['account_id']: r for r in report}

        # Account A: current balance 100,000 (one inflow); opening = current - inflow + 0 = 0.
        acc_a = by_id[self.account_a.id]
        self.assertEqual(acc_a['opening_balance'], Decimal('0.00'))
        self.assertEqual(acc_a['closing_balance'], Decimal('100000.00'))
        self.assertEqual(len(acc_a['rows']), 1)

        # Account B: current balance -5000 (one outflow); opening = current - 0 + outflow = 0.
        acc_b = by_id[self.account_b.id]
        self.assertEqual(acc_b['opening_balance'], Decimal('0.00'))
        self.assertEqual(acc_b['closing_balance'], Decimal('-5000.00'))
        self.assertEqual(len(acc_b['rows']), 1)

    def test_excluded_transaction_changes_opening_and_closing_balance(self):
        report = build_statement_report(
            self.household.id, [self.account_a.id],
            date(2026, 1, 1), date(2026, 1, 31),
            excluded_transaction_ids={self.account_a.id: [self.tx_a.id]},
        )
        acc_a = report[0]
        # Excluding the only transaction: opening == closing == current balance, no rows.
        self.assertEqual(acc_a['opening_balance'], Decimal('100000.00'))
        self.assertEqual(acc_a['closing_balance'], Decimal('100000.00'))
        self.assertEqual(len(acc_a['rows']), 0)

    def test_opening_balance_override_is_honored(self):
        report = build_statement_report(
            self.household.id, [self.account_a.id],
            date(2026, 1, 1), date(2026, 1, 31),
            opening_balance_overrides={self.account_a.id: '50000.00'},
        )
        acc_a = report[0]
        self.assertEqual(acc_a['opening_balance'], Decimal('50000.00'))
        # closing = opening + inflow = 50000 + 100000
        self.assertEqual(acc_a['closing_balance'], Decimal('150000.00'))

    def test_render_statement_pdf_produces_nonempty_bytes(self):
        report = build_statement_report(
            self.household.id, [self.account_a.id], date(2026, 1, 1), date(2026, 1, 31),
        )
        meta = {'household_name': self.household.name, 'generated_on': '2026-07-01', 'start_date': '2026-01-01', 'end_date': '2026-01-31'}
        pdf_bytes = render_statement_pdf(report, meta)
        self.assertGreater(len(pdf_bytes), 0)
        self.assertTrue(pdf_bytes.startswith(b'%PDF'))

    def test_render_statement_xlsx_round_trips(self):
        report = build_statement_report(
            self.household.id, [self.account_a.id, self.account_b.id], date(2026, 1, 1), date(2026, 1, 31),
        )
        meta = {'household_name': self.household.name, 'generated_on': '2026-07-01', 'start_date': '2026-01-01', 'end_date': '2026-01-31'}
        xlsx_bytes = render_statement_xlsx(report, meta)
        self.assertGreater(len(xlsx_bytes), 0)

        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        self.assertEqual(len(wb.sheetnames), 2)
        self.assertIn('Uncle Fund', wb.sheetnames)
        self.assertIn('Personal Savings', wb.sheetnames)
